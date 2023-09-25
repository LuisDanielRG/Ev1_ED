def RegistroNotas():
    # Opción 1. Registrar una nota.
    # Lista para almacenar los servicios y costos de cada cliente.
    detalles_servicios = []

    # Solicita el nombre y valida que no quede vacío.
    while True:
        nombre_cliente = input("\nIngresa el nombre del cliente: ")
        if nombre_cliente.strip() == "":
            print("**** ¡ERROR! El nombre no puede quedar vacío. ****")
        else:
            break
        
    # Solicita el RFC del cliente, y valida que este en un formato valido.
    patron_rfc = r'^[A-ZÑ&]{3,4}[0-9]{6}[A-Z0-9]{3}$'

    while True:
        rfc_cliente = input("\nIngrese el RFC del cliente: ")
        rfc_cliente = rfc_cliente.strip().upper()

        if not re.match(patron_rfc, rfc_cliente):
            print("**** ¡ERROR! El RFC no cumple con el patrón asignado. ****")
        else:
            letras_iniciales = rfc_cliente[:4] if len(rfc_cliente) == 13 else rfc_cliente[:3]
            fecha_str = rfc_cliente[len(letras_iniciales):len(letras_iniciales) + 6]

            try:
                fecha_rfc = datetime.datetime.strptime(fecha_str, "%y%m%d").date()
                if fecha_rfc > datetime.date.today():
                    print("**** ¡ERROR! El RFC no es válido. ****")
                    continue
                else:
                    if len(rfc_cliente) == 13:
                        print("El RFC ingresado es de persona física.")
                        break
                    else:
                        print("El RFC ingresado es de persona moral.")
                        break
            except Exception:
                print("**** ¡ERROR! La fecha no es válida. ****")
                continue

    # Solicita el correo electrónico del cliente y hace las respectivas validaciones.
    patron_correo = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

    while True:
        correo = input("\nIngrese el correo del cliente: ").lower()

        if not re.match(patron_correo,correo):
            print("**** ¡ERROR! El correo no cumple con el patrón asignado. ****")
            continue
        else:
            dominio = correo.split('@')[1]
            try:
                socket.gethostbyname(dominio)
                break
            except Exception:
                print("**** ¡ERROR! El dominio no existe, intenta de nuevo. ****")
                continue

    # Solicita la fecha y si esta en el formato correcto, la convierte a tipo de dato date.
    while True:
        fecha_capturada = input("\nIngresa la fecha (dd/mm/yyyy): ")
        try:
            fecha_procesada = datetime.datetime.strptime(fecha_capturada,"%d/%m/%Y").date()
        except Exception:
            print("**** ¡ERROR! Fecha inválida. Asegurate de seguir el formato dd/mm/yyyy. ****")
        else:
            # Una vez que la fecha se haya convertido, valida que no sea una fecha futura.
            if fecha_procesada < datetime.date.today():
                break
            else:
                print("**** ¡ERROR! La fecha ingresada es posterior a la fecha actual del sistema, por favor ingrese una fecha válida. ****")   
                continue

    # Pide el nombre y costo de los servicios, ambas no se pueden omitir.
    while True:
        servicios = input("\nIngrese el nombre del servicio: ")
        if servicios.strip() == "":
            print("**** ¡ERROR! El nombre del servicio no puede quedar vacío. Intenta de nuevo. *****")
            continue

        while True:
            costo = input("\nIngrese el costo del servicio: ")
                
            # Revisa que el dato ingresado coincida con el patrón, si coincide, entonces lo convierte a tipo de dato float, después de eso valida que sea mayor a 0.
            if re.match(r'^\d+(\.\d{0,2})?$', costo):
                costo_servicio = float(costo)
                if costo_servicio > 0:
                    break
                else:
                    print("**** ¡ERROR! El costo de servicio debe ser mayor a 0. Intenta de nuevo. ****")
                    continue  
            else:
                print("**** ¡ERROR! El costo de servicio no cumple con el formato, puede tener máximo 2 decimales o ninguno. ****")

        # Agrega los servicios y el costo a la lista detalles_servicios.
        detalles_servicios.append({"servicios": servicios, "costo_servicio": costo_servicio})

        # Pregunta si desea seguir agregando servicios.
        respuesta = input("\n¿Deseas seguir agregando servicios [SI / NO]? >> ")
        if respuesta.upper() != "SI":
            break
        
    # Realiza la suma de los cotos de los servicios de cada cliente.
    monto_a_pagar = sum(detalle["costo_servicio"] for detalle in detalles_servicios)

    # Genera el folio.
    folio = len(notas) + 1
        
    # Todos los datos solicitados y generados se agregan al diccionario notas.
    notas[folio] = {"fecha": fecha_procesada, "cliente": nombre_cliente, "rfc": rfc_cliente, "correo": correo, "detalle": detalles_servicios, "monto_a_pagar": monto_a_pagar}
    print(f"\n***** Nota registrada con éxito. Folio: {folio} *****")
    print("\nVolviendo al menú principal...")

def ConsultasReportes():
    # Opción 2. Consultas y Reportes.

    def ConsultaPorPeriodo():
        while True:
            # Solicita la fecha de inicio y si esta en el formato correcto, la convierte a tipo de dato date.
            fecha_capturadaInicial = input("\nIngresa la fecha de inicio (dd/mm/yyyy) o presiona Enter para usar la fecha predefinida (01/01/2000): ")
            if fecha_capturadaInicial == "":
                fecha_inicial = datetime.date(2000, 1, 1)
                print("Se usó la fecha predefinida (01/01/2000)")
            else:
                try:
                    fecha_inicial = datetime.datetime.strptime(fecha_capturadaInicial, "%d/%m/%Y").date()
                except ValueError:
                    print("***** ¡ERROR! Fecha inválida. Asegúrate de seguir el formato dd/mm/yyyy. *****")
                    continue

            # Solicita la fecha final y si está en el formato correcto, la convierte a tipo de dato date.
            fecha_capturadaFinal = input("\nIngresa la fecha final (dd/mm/yyyy) o presiona Enter para usar la fecha actual: ")
            if fecha_capturadaFinal == "":
                fecha_final = datetime.date.today()
                print(f"Usando la fecha actual ({fecha_final.strftime('%d/%m/%Y')})")
            else:
                try:
                    fecha_final = datetime.datetime.strptime(fecha_capturadaFinal, "%d/%m/%Y").date()
                except ValueError:
                    print("***** ¡ERROR! Fecha inválida. Asegúrate de seguir el formato dd/mm/yyyy. *****")
                    continue

            # Comprueba si la fecha final es igual o posterior a la fecha inicial.
            if fecha_final < fecha_inicial:
                print("***** ¡ERROR! La fecha final no puede ser anterior a la fecha inicial. *****")
            else:
                break

        # Filtra las notas por período.
        notas_periodo = {folio: nota for folio, nota in notas.items() if fecha_inicial <= nota['fecha'] <= fecha_final}

        if notas_periodo:
            datos_por_periodo = []

            total_monto = 0
            for folio, nota in notas_periodo.items():
                fecha = nota['fecha']
                monto = nota['monto_a_pagar']
                total_monto += monto

                datos_por_periodo.append([folio, fecha.strftime("%d/%m/%Y"), nota['cliente'], nota['rfc'], nota['correo'], f"${monto:.2f}"])

            promedio_monto = total_monto / len(notas_periodo)

            # Muestra de forma tabular las notas encontradas en el período solicitado.
            print("\nREPORTE DE NOTAS POR PERÍODO:")
            tabla_por_periodo = tabulate(datos_por_periodo, headers=["Folio", "Fecha", "Nombre del Cliente", "RFC del Cliente", "Correo del Cliente", "Monto a Pagar"], tablefmt="grid")
            print(tabla_por_periodo)

            print(f"\nMonto Promedio: ${promedio_monto:.2f}")
        else:
            print("\n***** No hay notas emitidas para el período especificado. *****")


    def ConsultaPorFolio():
         
        
        # Opción 2.2 Consulta por folio.
        while True:
            # Solicita el folio a consultar.
            folio = input("\nIngrese el folio de la nota a consultar: ")

            try:
                # Convierte el folio a tipo de dato int.
                folio_a_consultar = int(folio)
                break
            except Exception:
                print("***** ¡ERROR! El folio debe ser un número entero. *****")

        # Revisa que el folio ingresado exista.
        if folio_a_consultar in notas:
            nota = notas[folio_a_consultar]
            fecha = nota['fecha']

            # Lista para la nota con el folio solicitado.
            datos_por_folio = []

            for detalle in nota['detalle']:
                if not datos_por_folio:
                    datos_por_folio.append([folio_a_consultar, fecha.strftime("%d/%m/%Y"), nota['cliente'], nota['rfc'], nota['correo'], detalle['servicios'], f"${detalle['costo_servicio']:.2f}", f"${nota['monto_a_pagar']:.2f}"])
                else:
                    datos_por_folio.append(["", "", "", "", "", detalle['servicios'], f"${detalle['costo_servicio']:.2f}", ""])
            
            # Muestra la nota en forma de tabla tabular.
            tabla_por_folio = tabulate(datos_por_folio, headers=["Folio", "Fecha", "Nombre del Cliente", "RFC del Cliente", "Correo del Cliente", "Servicio(s)", "Costo(s)", "Monto a Pagar"], tablefmt="grid")

            print("\nREPORTE DE NOTAS POR FOLIO:")
            print(tabla_por_folio)
        else:
            print("***** La nota no se encuentra en el sistema. *****")
            
    def ConsultaPorCliente():
        # Opción 2.3 Consulta por cliente.
        rfc_unicos = set() # Conjunto para almacenar RFC únicos, sin duplicidades.

        for folio, nota in notas.items():
            rfc = nota['rfc']
            rfc_unicos.add(rfc)
        
        lista_rfc = sorted(list(rfc_unicos)) # Convierte el conjunto a una lista

        while True:
            print("\nRFC Disponbles para consultar:")
            for i, rfc in enumerate(lista_rfc, start=1):
                print(f"{i}. RFC: {rfc}")
            
            seleccion_rfc = input("\nIngresa el número correspondiente al RFC que deseas consultar: ")

            try:
                seleccion = int(seleccion_rfc)
                if 1 <= seleccion <= len(lista_rfc):
                    rfc_a_consultar = lista_rfc[seleccion - 1]
                    notas_cliente = [(folio, nota) for folio, nota in notas.items() if nota['rfc'] == rfc_a_consultar]
                    notas_cliente_ordenadas = sorted(notas_cliente, key=lambda x: x[0])

                    print(f"\nNOTAS PARA EL CLIENTE CON EL RFC {rfc_a_consultar}")
                    montos_consultados = []

                    for folio, nota in notas_cliente_ordenadas:
                        fecha = nota['fecha']
                        montos_consultados.append(nota['monto_a_pagar'])

                        print("\n\tDETALLE DE LAS NOTAS:")
                        print("\t","- "*20)
                        print(f"\t\tFolio                 : {folio}")
                        print(f"\t\tFecha                 : {fecha.strftime('%d/%m/%Y')}")
                        print(f"\t\tMonto a Pagar         : ${nota['monto_a_pagar']}")
                        print("\t\tServicios:")
                        for detalle in nota['detalle']:
                            print(f"\t\t- Servicio            : {detalle['servicios']}")
                            print(f"\t\t  Costo               : ${detalle['costo_servicio']}")
                    
                    montos_consultados_np = np.array(montos_consultados)
                    prom_montos = np.mean(montos_consultados_np)
                    print(f"\nPromedio de los montos de las notas consultadas:  ${prom_montos:.2f}")
                    break
                else:
                    print("**** ¡ERROR! Selección fue del rango. Ingresa un número válido.")
            except Exception:
                print("**** ¡ERROR! Dato no válido, debe ser un número. ****")

        while True:
            # Pregunta si desea exportar la información a un archivo de excel.
            exportar_a_excel = input("\n¿Deseas exportar la información anterior a un archivo de excel [SI / NO] >> ")
            if exportar_a_excel.upper() == "SI":
                fecha_actual = datetime.date.today().strftime("%d-%m-%Y")
                
                libro_excel = openpyxl.Workbook()
                hoja_trabajo = libro_excel.active

                hoja_trabajo.append(["Folio", "Fecha", "Nombre del Cliente", "RFC del Cliente", "Correo del Cliente", "Monto a Pagar", "Detalle de Servicios"])

                for folio, nota in notas_cliente_ordenadas:
                    fecha = nota['fecha'].strftime("%d/%m/%Y")
                    detalle_servicios = ""
                    for detalle in nota['detalle']:
                        detalle_servicios += f"{detalle['servicios']}: ${detalle['costo_servicio']}\n"
                    hoja_trabajo.append([folio, fecha, nota['cliente'], nota['rfc'], nota['correo'], nota['monto_a_pagar'], detalle_servicios])

                for row in hoja_trabajo.iter_rows(min_row=2, max_row=hoja_trabajo.max_row, min_col=7, max_col=7):
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                
                archivo_excel = f"{rfc_a_consultar}_{fecha_actual}.xlsx"
                libro_excel.save(archivo_excel)
                print(f"La información ha sido exportada con el nombre {archivo_excel}")
                break
            elif exportar_a_excel.upper() == "NO":
                print("**** La información no fue exportada. ****")
                break
            else:
                print("**** ¡ERROR! Ingresa una respuesta válida [SI / NO]. ****")
                continue
                        
    
    while True:
            # Submenú de consultas y reportes.
            print("\n************************************")
            print("    Menú de Consultas y Reportes    ")
            print("************************************")
            print("1. Consulta por período")
            print("2. Consulta por folio")
            print("3. Consulta por cliente")
            print("4. Volver al menú principal")
            print("************************************")

            sub_opcion_capturada = input("\nSeleccione una opción: ")

            if sub_opcion_capturada.strip() == "":
                print("**** ¡ERROR! No puede quedarse vacío. ****")
                continue
            else:
                try:
                    # Convierte la opcion capturada a tipo de dato int.
                    sub_opcion = int(sub_opcion_capturada)
                except Exception:
                    print("***** ¡ERROR! La opción debe ser un número entero. *****")
                    continue

                if sub_opcion == 1:
                    ConsultaPorPeriodo()
                elif sub_opcion == 2:
                    ConsultaPorFolio()
                elif sub_opcion == 3:
                    ConsultaPorCliente()
                elif sub_opcion == 4:
                    print("Volviendo al menú principal...")
                    break
                else:
                    print("***** Opción no válida. *****")
        
        
        
        
def CancelarNota():
    # Opción 3. Cancelar notas.
    while True:
        # Solicitar folio a cancelar.
        folio = input("\nIngrese el folio de la nota que desea cancelar: ")
        
        try:
            # Convierte a tipo de dato int.
            folio_a_cancelar = int(folio)
            break
        except Exception:
            print("***** ¡ERROR! El folio debe ser un número entero. *****")

    # Busca la nota con el folio ingresado.
    if folio_a_cancelar in notas:
        nota = notas[folio_a_cancelar]

        fecha_objeto = nota['fecha']
        fecha = fecha_objeto.strftime("%d/%m/%Y")

        # Muestra los datos de la nota antes de cancelar.
        print("\n\tDETALLE DE LA NOTA A CANCELAR:")
        print("\t","- "*20)
        print(f"\t\tFolio                 : {folio_a_cancelar}")
        print(f"\t\tFecha                 : {fecha}")
        print(f"\t\tNombre del Cliente    : {nota['cliente']}")
        print(f"\t\tRFC del Cliente       : {nota['rfc']}")
        print(f"\t\tCorreo del Cliente    : {nota['correo']}")
        print(f"\t\tMonto a Pagar         : ${nota['monto_a_pagar']}")
        print("\t\tServicios:")
        for detalle in nota['detalle']:
            print(f"\t\t- Servicio            : {detalle['servicios']}")
            print(f"\t\t  Costo               : ${detalle['costo_servicio']}")
    
        while True:
            # Se pide la confirmación de la cancelación.
            confirmacion = input("\n¿Estas seguro de que deseas cancelar la nota [SI / NO]? >> ")
            if confirmacion.upper() == "SI":
                # Se cancela.
                notas_canceladas[folio_a_cancelar] = nota
                del notas[folio_a_cancelar]
                print(f"***** La nota con el folio {folio_a_cancelar} ha sido cancelada exitosamente. *****")
                break
            elif confirmacion.upper() == "NO":
                print(f"***** La nota con el folio {folio_a_cancelar} no fue cancelada. *****")
                break
            else:
                print("***** ¡ERROR! Ingresa una respuesta válida [SI / NO]. *****")
                continue
    else:
        print("***** El folio ingresado no existe en el sistema. *****")

def RecuperarNota():
    # Opción 4. Recuperar notas.
    if notas_canceladas:
        # Muestra las notas que han sido canceladas anteriormente.
        datos_cancelados = []

        for folio, nota in notas_canceladas.items():
            fecha = nota['fecha']

            datos_cancelados.append([folio, fecha.strftime("%d/%m/%Y"), nota['cliente'], nota['rfc'], nota['correo'], f"${nota['monto_a_pagar']:.2f}"])
            
        print("\nNOTAS CANCELADAS:")
        tabla_notas_canceladas = tabulate(datos_cancelados, headers=["Folio", "Fecha", "Nombre del Cliente", "RFC del Cliente", "Correo del Cliente", "Monto a Pagar"], tablefmt="grid")
        print(tabla_notas_canceladas)
    else:
        print("\n***** No hay notas canceladas hasta el momento. *****")
        return

    while True:
        # Solicita el folio de la nota a recuperar.
        folio = input("\nIngresa el folio de la que nota que deseas recuperar (Presiona ENTER si no deseas recuperar ninguna nota de las presentadas): ")
    
        if folio.strip() == "":
            # Regresa al menú principal si así se desea.
            print("Volviendo al menú principal...")
            return
        else:
            try:
                # Convierte el folio ingresado a tipo de dato int.
                folio_a_recuperar = int(folio)
                break
            except Exception:
                print("***** ¡ERROR! El folio debe ser un número entero. *****")

    # Busca el folio de la nota en las notas canceladas.
    if folio_a_recuperar in notas_canceladas:
        nota = notas_canceladas[folio_a_recuperar]
       
        # Muestra los detalles (servicio y costo de servicio) de la nota.
        print("\n\tDETALLE DE LA NOTA:")
        print("\t","- "*20)
        for detalle in nota['detalle']:
            print(f"\t\t- Servicio: {detalle['servicios']}")
            print(f"\t\t  Costo: ${detalle['costo_servicio']}")

        while True:
            # Se confirma si se desea recuperar la nota o no.
            confirmacion = input("\n¿Esta seguro de recuperar esta nota [SI / NO]? >> ")
            if confirmacion.upper() == "SI":
                # Recupera la nota.
                notas[folio_a_recuperar] = nota
                del notas_canceladas[folio_a_recuperar]
                print(f"***** La nota con el folio {folio_a_recuperar} ha sido recuperada exitosamente. *****")
                break
            elif confirmacion.upper() == "NO":
                print(f"***** La nota con el folio {folio_a_recuperar} no fue recuperada. *****")
                break
            else:
                print("***** ¡ERROR! Ingresa una respuesta válida [SI / NO]. *****")
                continue
    else:
        print("***** El folio ingresado no corresponde a una nota cancelada. *****")

def SalidaSistema():
    # Opción 5. Salida del sistema.
    while True:
        confirmacion_salida = input("\n¿Estas seguro de salir del sistema [SI / NO]? >> ")
        if confirmacion_salida.upper() == "SI":
            print("Gracias por visitar este sistema. ¡Vuelva pronto!\n")
            exit()
        elif confirmacion_salida.upper() == "NO":
            print("Volviendo al menú principal...")
            return
        else:
            print("**** ¡ERROR! Ingresa una respuesta válida [SI / NO]. ****")
            continue

# Menú Principal
import datetime
import re
import socket
from tabulate import tabulate
import openpyxl
import numpy as np
import csv


notas = {}
notas_canceladas = {}

print("\n  ¡BIENVENIDO/A AL SISTEMA DEL TALLER MECÁNICO!  ")
while True:
    print("\n*************************************************")
    print("        ¿Qué desea realizar el día de hoy?       ")
    print("*************************************************")
    print("(1) Registrar una nota")
    print("(2) Consultas y reportes")
    print("(3) Cancelar una nota")
    print("(4) Recuperar una nota")
    print("(5) Salir del sistema")
    print("*************************************************")

    opcion_capturada = input("\nSeleccione una opción: ")

    if opcion_capturada.strip() == "":
        print("**** ¡ERROR! No puede quedarse vacío. ****")
        continue
    else:
        try:
            # Convierte la opcion capturada a tipo de dato int.
            opcion = int(opcion_capturada)
        except Exception:
            print("***** ¡ERROR! La opción debe ser un número entero. *****")
            continue

        if opcion == 1:
            RegistroNotas()
        elif opcion == 2:
            ConsultasReportes()
        elif opcion == 3:
            CancelarNota()
        elif opcion == 4:
            RecuperarNota()
        elif opcion == 5:
            SalidaSistema()
        else:
            print("**** Opción no válida. ****")
